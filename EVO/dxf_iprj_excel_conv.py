import os
import base64
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import load_workbook
import ezdxf  # For reading and processing DXF files
import numpy as np  # For numerical operations and comparisons

def sanitize_dataframe(df):
    return df.applymap(lambda x: str(x) if isinstance(x, (dict, list)) else ("" if x is None else x))

def iprj_to_excel(iprj_path, output_excel_path):
    tree = ET.parse(iprj_path)
    root = tree.getroot()

    sensor_data = {}
    background_data = {}
    lineals = []
    textlabels = []

    for elem in root.findall('.//Configuration'):
        for key, value in elem.attrib.items():
            if key.startswith("Radarsensor_"):
                parts = key.split('_')
                if parts[1].isdigit():
                    sensor_idx = int(parts[1])
                    if sensor_idx not in sensor_data:
                        sensor_data[sensor_idx] = {
                            "positional": {},
                            "eventzones": [],
                            "ignorezones": [],
                            "conditions": []
                        }

                    section = parts[2] if len(parts) > 2 else None

                    if section in ['Position', 'AzimuthAngle', 'ElevationAngle', 'RoadGradientAngle',
                                   'InstallationHeight', 'RainInterferenceThreshold', 'HighwayMode',
                                   'MaxStopTime', 'FrequencyChannel', 'Gps']:
                        field_name = "_".join(parts[2:])
                        sensor_data[sensor_idx]["positional"][field_name] = value

                    elif section == "EventZone":
                        zone_idx = int(parts[3])
                        key_name = "_".join(parts[4:])
                        while len(sensor_data[sensor_idx]["eventzones"]) <= zone_idx:
                            sensor_data[sensor_idx]["eventzones"].append({"Index": zone_idx})
                        sensor_data[sensor_idx]["eventzones"][zone_idx][key_name] = value

                    elif section == "IgnoreZone":
                        zone_idx = int(parts[3])
                        key_name = "_".join(parts[4:])
                        while len(sensor_data[sensor_idx]["ignorezones"]) <= zone_idx:
                            sensor_data[sensor_idx]["ignorezones"].append({"Index": zone_idx})
                        sensor_data[sensor_idx]["ignorezones"][zone_idx][key_name] = value

                    elif section == "Condition":
                        continue  # We get conditions via EventZone nesting instead

                elif parts[1] == "nrOfSensors":
                    continue  # Already inferred from keys

            elif key.startswith("Lineals_"):
                parts = key.split('_')
                idx = int(parts[1])
                while len(lineals) <= idx:
                    lineals.append({"Index": idx})
                subkey = "_".join(parts[2:])
                lineals[idx][subkey] = value

            elif key.startswith("Textlabel_"):
                parts = key.split('_')
                idx = int(parts[1])
                while len(textlabels) <= idx:
                    textlabels.append({"Index": idx})
                subkey = "_".join(parts[2:])
                textlabels[idx][subkey] = value

            elif key == "BackgroundImage":
                background_data["image_base64"] = value
            elif key.startswith("Background"):
                background_data[key] = value
            elif key in ['ReferenceLength', 'MeterPerPixel',
                         'MeterReference0_X', 'MeterReference0_Y',
                         'MeterReference1_X', 'MeterReference1_Y']:
                background_data[key] = value

    # Extract conditions from eventzones
    for sensor_idx, sensor in sensor_data.items():
        conditions = []
        for zone in sensor["eventzones"]:
            zone_idx = zone["Index"]
            condition_keys = [k for k in zone.keys() if k.startswith("Condition_")]
            
            # Group condition keys by their index
            condition_groups = {}
            for k in condition_keys:
                parts = k.split('_')
                cond_idx = int(parts[1])
                if cond_idx not in condition_groups:
                    condition_groups[cond_idx] = []
                condition_groups[cond_idx].append(k)
            
            # Create condition entries
            for cond_idx, keys in condition_groups.items():
                condition = {"EventZone": zone_idx}
                for k in keys:
                    short_key = k.replace(f"Condition_{cond_idx}_", "")
                    condition[short_key] = zone[k]
                
                # Add condition to the list (we'll filter by Enable later)
                if keys:  # Only add if there are actually condition keys
                    conditions.append(condition)
        
        sensor["conditions"] = conditions

    # Start writing Excel
    wb = Workbook()
    wb.remove(wb.active)

    # Function to add a section title
    def add_section_title(worksheet, row, title):
        cell = worksheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True)
        return row + 1

    for sensor_idx, sensor in sensor_data.items():
        ws = wb.create_sheet(title=f"Sensor {sensor_idx + 1}")
        current_row = 1

        # Write positional data vertically
        pos_items = list(sensor["positional"].items())
        for i, (key, val) in enumerate(pos_items):
            ws[f"A{current_row + i}"] = key
            ws[f"B{current_row + i}"] = val
        
        current_row += len(pos_items) + 2  # Add empty row

        # Write Event Zones section
        current_row = add_section_title(ws, current_row, "Event Zones")

        # Filter only enabled event zones
        enabled_zones = [zone for zone in sensor["eventzones"] 
                        if zone.get("Enable", "0") in ("1", 1)]

        # Adjust the "EventZone" column to start at 1 for Excel
        for zone in enabled_zones:
            zone["EventZone"] = zone["Index"] + 1  # Add 1 to make it 1-based for Excel

        # Remove condition fields from zones
        df_zones_raw = pd.DataFrame(enabled_zones)
        condition_cols = [col for col in df_zones_raw.columns if col.startswith("Condition_")] if not df_zones_raw.empty else []
        df_zones = df_zones_raw.drop(columns=condition_cols) if not df_zones_raw.empty else pd.DataFrame()

        # Even if empty, create a DataFrame with the expected columns
        if df_zones.empty:
            df_zones = pd.DataFrame(columns=["EventZone", "Enable", "ZonePoint_0_X", "ZonePoint_0_Y", 
                                            "ZonePoint_1_X", "ZonePoint_1_Y", "ZonePoint_2_X", "ZonePoint_2_Y",
                                            "ZonePoint_3_X", "ZonePoint_3_Y"])
        else:
            # Reorder columns to make "EventZone" the left-most column and exclude "Index"
            columns = ["EventZone"] + [col for col in df_zones.columns if col != "Index" and col != "EventZone"]
            df_zones = df_zones[columns]

        df_zones = sanitize_dataframe(df_zones)

        for r_idx, row in enumerate(dataframe_to_rows(df_zones, index=False, header=True)):
            for c_idx, value in enumerate(row):
                ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)

        current_row += len(df_zones) + 2  # Include header row and add empty row

        # Write Conditions section
        current_row = add_section_title(ws, current_row, "Conditions")

        # Filter only enabled conditions
        enabled_conditions = [cond for cond in sensor["conditions"] 
                            if cond.get("Enable", "0") in ("1", 1, "True", True, "true", "yes", "YES")]

        # Adjust the "EventZone" column to use 1-based indexing
        for cond in enabled_conditions:
            cond["EventZone"] = cond["EventZone"] + 1  # Convert 0-based to 1-based indexing

        df_conditions = pd.DataFrame(enabled_conditions)
        if df_conditions.empty:
            df_conditions = pd.DataFrame(columns=["EventZone", "Enable", "OutputNumber", "ConditionClass",
                                                "Direction", "VelocityMin", "VelocityMax", "QueuelengthMin",
                                                "QueuelengthMax", "EventMessageDelay", "EventMessageExtend",
                                                "EtaMin", "EtaMax"])

        df_conditions = sanitize_dataframe(df_conditions)

        for r_idx, row in enumerate(dataframe_to_rows(df_conditions, index=False, header=True)):
            for c_idx, value in enumerate(row):
                ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)

        current_row += len(df_conditions) + 2  # Include header row and add empty row

        # Write Ignore Zones section
        current_row = add_section_title(ws, current_row, "Ignore Zones")
        
        # Filter only enabled ignore zones
        enabled_ignore_zones = [zone for zone in sensor["ignorezones"] 
                              if zone.get("Enable", "0") in ("1", 1)]
        
        df_ignores = pd.DataFrame(enabled_ignore_zones)
        if df_ignores.empty:
            df_ignores = pd.DataFrame(columns=["Index", "Enable", "ZonePoint_0_X", "ZonePoint_0_Y", 
                                            "ZonePoint_1_X", "ZonePoint_1_Y", "ZonePoint_2_X", "ZonePoint_2_Y",
                                            "ZonePoint_3_X", "ZonePoint_3_Y"])
        
        df_ignores = sanitize_dataframe(df_ignores)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_ignores, index=False, header=True)):
            for c_idx, value in enumerate(row):
                ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)

    # Combined Annotations sheet (Lineals and Textlabels)
    ws_annotations = wb.create_sheet("Annotations")
    current_row = 1
    
    # Add Lineals section
    current_row = add_section_title(ws_annotations, current_row, "Lineals")
    
    # Filter only enabled lineals
    enabled_lineals = [l for l in lineals if l.get("Enable", "0") in ("1", 1)]
    
    df_lineals = pd.DataFrame(enabled_lineals)
    if df_lineals.empty:
        df_lineals = pd.DataFrame(columns=["Index", "Type", "Enable", "Point_0_X", "Point_0_Y", "Point_1_X", "Point_1_Y"])
    else:
        df_lineals["Type"] = "Lineal"  # Add Type column
    
    df_lineals = sanitize_dataframe(df_lineals)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_lineals, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws_annotations.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
    
    current_row += len(df_lineals) + 2  # Include header row and add empty row
    
    # Add Textlabels section
    current_row = add_section_title(ws_annotations, current_row, "Textlabels")
    
    # Filter only enabled textlabels
    enabled_textlabels = [t for t in textlabels if t.get("Enable", "0") in ("1", 1, "True", True, "true", "yes", "YES")]
    
    df_textlabels = pd.DataFrame(enabled_textlabels)
    if df_textlabels.empty:
        df_textlabels = pd.DataFrame(columns=["Index", "Type", "Enable", "Text", "Position_X", "Position_Y", "FontSize",
                                           "FontBold", "FontUnderline", "FontItalic", "RotationAngle",
                                           "Textcolor_Red", "Textcolor_Green", "Textcolor_Blue"])
    else:
        df_textlabels["Type"] = "Textlabel"  # Add Type column
    
    df_textlabels = sanitize_dataframe(df_textlabels)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_textlabels, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws_annotations.cell(row=current_row + r_idx, column=c_idx + 1, value=value)

    # Background sheet
    if background_data:
        ws_bg = wb.create_sheet("BackgroundImageInfo")

        image_path = os.path.splitext(output_excel_path)[0] + "_background.png"
        if "image_base64" in background_data:
            with open(image_path, "wb") as img_file:
                img_file.write(base64.b64decode(background_data["image_base64"]))
            
            del background_data["image_base64"]
            background_data["image_path"] = image_path

        for i, (key, val) in enumerate(background_data.items()):
            ws_bg[f"A{i+1}"] = key
            ws_bg[f"B{i+1}"] = val

    wb.save(output_excel_path)

def excel_to_iprj(excel_path, output_iprj_path):
    """
    Convert Excel file to IPRJ format (reverse of parse_iprj_to_excel)
    
    Args:
        excel_path: Path to the Excel file
        output_iprj_path: Path to save the output IPRJ file
    """
    # Load the Excel workbook
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    
    # Create XML root element
    root = ET.Element("IwarrSensorProject")
    config = ET.SubElement(root, "Configuration")
    
    # Dictionary to store all configuration attributes
    config_attrs = {}
    
    # Find sensor sheets and infer number of sensors
    sensor_sheets = [s for s in sheets if s.startswith("Sensor ")]
    num_sensors = len(sensor_sheets)
    config_attrs["Radarsensor_nrOfSensors"] = str(num_sensors)
    
    # Process each sensor sheet
    for sheet_name in sensor_sheets:
        ws = wb[sheet_name]
        
        # Extract sensor index (0-based)
        sensor_idx = int(sheet_name.split(" ")[1]) - 1
        
        # Process positional data (vertical key-value pairs)
        row = 1
        while row <= ws.max_row and ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=1).value != "Event Zones":
            key = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if value is not None:  # Skip empty values
                config_attrs[f"Radarsensor_{sensor_idx}_{key}"] = str(value)
            row += 1
        
        # Skip to Event Zones section if it exists
        while row <= ws.max_row and (ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value != "Event Zones"):
            row += 1
        
        # Process Event Zones if we found the section
        if row <= ws.max_row and ws.cell(row=row, column=1).value == "Event Zones":
            row += 1  # Skip section header

            # Get headers for event zones
            headers = []
            col = 1
            while col <= ws.max_column and ws.cell(row=row, column=col).value is not None:
                headers.append(ws.cell(row=row, column=col).value)
                col += 1

            if headers:  # Only proceed if headers were found
                row += 1  # Move to first data row

                # Process each event zone
                while row <= ws.max_row and ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=1).value != "Conditions":
                    # Map "EventZone" column back to 0-based indexing
                    event_zone_idx = None
                    for i, header in enumerate(headers):
                        if header == "EventZone":
                            event_zone_idx = i
                            break

                    if event_zone_idx is not None:
                        zone_value = ws.cell(row=row, column=event_zone_idx + 1).value
                        if zone_value is not None:
                            zone_idx = int(zone_value) - 1  # Subtract 1 to make it 0-based
                        else:
                            zone_idx = row - 3  # Default if cell is empty
                    else:
                        # Fallback to using the first column or row number
                        first_col_value = ws.cell(row=row, column=1).value
                        if first_col_value is not None and (isinstance(first_col_value, (int, float)) or (isinstance(first_col_value, str) and first_col_value.isdigit())):
                            zone_idx = int(first_col_value) - 1  # Subtract 1 to make it 0-based
                        else:
                            # Use row number as index (offset by headers and section title)
                            zone_idx = row - 3

                    for col_idx, header in enumerate(headers):
                        value = ws.cell(row=row, column=col_idx + 1).value
                        if value is not None and value != "":  # Skip empty values
                            if header == "ZoneName":
                                # Modify ZoneName to include EventZone
                                value = f"{zone_idx + 1}: {value}"
                            config_attrs[f"Radarsensor_{sensor_idx}_EventZone_{zone_idx}_{header}"] = str(value)

                    row += 1
                    if row > ws.max_row or ws.cell(row=row, column=1).value == "Conditions":
                        break
        
        # Skip to Conditions section if it exists
        while row <= ws.max_row and (ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value != "Conditions"):
            row += 1
        
        # Process Conditions if we found the section
        if row <= ws.max_row and ws.cell(row=row, column=1).value == "Conditions":
            row += 1  # Skip section header
            
            # Get headers for conditions - first check if there are any headers
            if row <= ws.max_row:
                cond_headers = []
                col = 1
                while col <= ws.max_column and ws.cell(row=row, column=col).value is not None:
                    cond_headers.append(ws.cell(row=row, column=col).value)
                    col += 1
                
                if cond_headers:  # Only proceed if headers were found
                    row += 1  # Move to first data row
                    
                    # Check if EventZone column exists
                    event_zone_col = None
                    for i, header in enumerate(cond_headers):
                        if header == "EventZone":
                            event_zone_col = i
                            break
                    
                    # Process each condition
                    cond_counter = {}  # Track condition indexes per event zone
                    while row <= ws.max_row and ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=1).value != "Ignore Zones":
                        # Get the event zone index - this must exist for conditions
                        if event_zone_col is not None and ws.cell(row=row, column=event_zone_col + 1).value is not None:
                            event_zone_idx = int(ws.cell(row=row, column=event_zone_col + 1).value) - 1  # Convert 1-based to 0-based indexing
                        else:
                            # If EventZone column is missing or empty, use a default
                            event_zone_idx = 0  # Default to first event zone
                        
                        # Assign condition index based on event zone
                        if event_zone_idx not in cond_counter:
                            cond_counter[event_zone_idx] = 0
                        condition_idx = cond_counter[event_zone_idx]
                        cond_counter[event_zone_idx] += 1
                        
                        for col_idx, header in enumerate(cond_headers):
                            value = ws.cell(row=row, column=col_idx + 1).value
                            if value is not None and value != "":  # Skip empty values
                                if header != "EventZone":  # Don't include EventZone as an attribute
                                    config_attrs[f"Radarsensor_{sensor_idx}_EventZone_{event_zone_idx}_Condition_{condition_idx}_{header}"] = str(value)
                        
                        row += 1
                        if row > ws.max_row or ws.cell(row=row, column=1).value == "Ignore Zones":
                            break
        
        # Skip to Ignore Zones section if it exists
        while row <= ws.max_row and (ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value != "Ignore Zones"):
            row += 1
        
        # Process Ignore Zones if we found the section
        if row <= ws.max_row and ws.cell(row=row, column=1).value == "Ignore Zones":
            row += 1  # Skip section header
            
            # Get headers for ignore zones
            ignore_headers = []
            col = 1
            while col <= ws.max_column and ws.cell(row=row, column=col).value is not None:
                ignore_headers.append(ws.cell(row=row, column=col).value)
                col += 1
            
            if ignore_headers:  # Only proceed if headers were found
                row += 1  # Move to first data row
                
                # Process each ignore zone
                while row <= ws.max_row and ws.cell(row=row, column=1).value is not None:
                    # Find the Index column
                    index_col = None
                    for i, header in enumerate(ignore_headers):
                        if header == "Index":
                            index_col = i
                            break
                    
                    # If Index column found, use it, otherwise use the row number as index
                    if index_col is not None and ws.cell(row=row, column=index_col + 1).value is not None:
                        zone_idx = int(ws.cell(row=row, column=index_col + 1).value)
                    else:
                        # Fallback to using the first column or row number
                        first_col_value = ws.cell(row=row, column=1).value
                        if first_col_value is not None and (isinstance(first_col_value, (int, float)) or (isinstance(first_col_value, str) and first_col_value.isdigit())):
                            zone_idx = int(first_col_value)
                        else:
                            # Use row number as index (offset by headers and section title)
                            zone_idx = row - len(ignore_headers) - 2
                    
                    for col_idx, header in enumerate(ignore_headers):
                        value = ws.cell(row=row, column=col_idx + 1).value
                        if value is not None and value != "":  # Skip empty values
                            config_attrs[f"Radarsensor_{sensor_idx}_IgnoreZone_{zone_idx}_{header}"] = str(value)
                    
                    row += 1
    
    # Process Annotations sheet if it exists
    if "Annotations" in sheets:
        ws = wb["Annotations"]
        row = 1
        
        # Skip to Lineals section
        while row < ws.max_row and (ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value != "Lineals"):
            row += 1
        
        # Process Lineals if we found the section
        if row < ws.max_row and ws.cell(row=row, column=1).value == "Lineals":
            row += 1  # Skip section header
            
            # Get headers for lineals
            lineal_headers = []
            col = 1
            while col <= ws.max_column and ws.cell(row=row, column=col).value is not None:
                lineal_headers.append(ws.cell(row=row, column=col).value)
                col += 1
            
            if lineal_headers:  # Only proceed if headers were found
                row += 1  # Move to first data row
                
                # Process each lineal
                while row < ws.max_row and ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=1).value != "Textlabels":
                    # Find the Index column
                    index_col = None
                    for i, header in enumerate(lineal_headers):
                        if header == "Index":
                            index_col = i
                            break
                    
                    # If Index column found, use it, otherwise use the row number as index
                    if index_col is not None and ws.cell(row=row, column=index_col + 1).value is not None:
                        lineal_idx = int(ws.cell(row=row, column=index_col + 1).value)
                    else:
                        # Fallback to using the first column or row number
                        first_col_value = ws.cell(row=row, column=1).value
                        if first_col_value is not None and (isinstance(first_col_value, (int, float)) or (isinstance(first_col_value, str) and first_col_value.isdigit())):
                            lineal_idx = int(first_col_value)
                        else:
                            # Use row number as index (offset by headers and section title)
                            lineal_idx = row - len(lineal_headers) - 2
                    
                    for col_idx, header in enumerate(lineal_headers):
                        value = ws.cell(row=row, column=col_idx + 1).value
                        if value is not None and header != "Type" and value != "":  # Skip empty values and Type field
                            config_attrs[f"Lineals_{lineal_idx}_{header}"] = str(value)
                    
                    row += 1
                    if row >= ws.max_row or ws.cell(row=row, column=1).value == "Textlabels":
                        break
        
        # Skip to Textlabels section
        while row < ws.max_row and (ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value != "Textlabels"):
            row += 1
        
        # Process Textlabels if we found the section
        if row < ws.max_row and ws.cell(row=row, column=1).value == "Textlabels":
            row += 1  # Skip section header
            
            # Get headers for textlabels
            textlabel_headers = []
            col = 1
            while col <= ws.max_column and ws.cell(row=row, column=col).value is not None:
                textlabel_headers.append(ws.cell(row=row, column=col).value)
                col += 1
            
            if textlabel_headers:  # Only proceed if headers were found
                row += 1  # Move to first data row
                
                # Process each textlabel
                while row <= ws.max_row and ws.cell(row=row, column=1).value is not None:
                    # Find the Index column
                    index_col = None
                    for i, header in enumerate(textlabel_headers):
                        if header == "Index":
                            index_col = i
                            break
                    
                    # If Index column found, use it, otherwise use the row number as index
                    if index_col is not None and ws.cell(row=row, column=index_col + 1).value is not None:
                        textlabel_idx = int(ws.cell(row=row, column=index_col + 1).value)
                    else:
                        # Fallback to using the first column or row number
                        first_col_value = ws.cell(row=row, column=1).value
                        if first_col_value is not None and (isinstance(first_col_value, (int, float)) or (isinstance(first_col_value, str) and first_col_value.isdigit())):
                            textlabel_idx = int(first_col_value)
                        else:
                            # Use row number as index (offset by headers and section title)
                            textlabel_idx = row - len(textlabel_headers) - 2
                    
                    for col_idx, header in enumerate(textlabel_headers):
                        value = ws.cell(row=row, column=col_idx + 1).value
                        if value is not None and header != "Type" and value != "":  # Skip empty values and Type field
                            config_attrs[f"Textlabel_{textlabel_idx}_{header}"] = str(value)
                    
                    row += 1
    
    # Process BackgroundImageInfo sheet if it exists
    if "BackgroundImageInfo" in sheets:
        ws = wb["BackgroundImageInfo"]
        
        # Process background data
        for row in range(1, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            
            if key is not None and value is not None:
                # Handle image path specially
                if key == "image_path":
                    try:
                        # Check if file exists
                        if os.path.exists(value):
                            with open(value, "rb") as img_file:
                                img_data = img_file.read()
                                img_base64 = base64.b64encode(img_data).decode('utf-8')
                                config_attrs["BackgroundImage"] = img_base64
                        else:
                            print(f"Warning: Image file not found: {value}")
                    except Exception as e:
                        print(f"Warning: Could not read image file {value}: {e}")
                else:
                    config_attrs[key] = str(value)
    
    # Set all attributes on the Configuration element
    for key, value in config_attrs.items():
        config.set(key, value)
    
    # Generate XML string
    tree = ET.ElementTree(root)
    
    # Standard XML declaration
    xml_declaration = '<?xml version="1.0" encoding="utf-8"?>\n'
    
    # Convert to string and write to file
    with open(output_iprj_path, "w", encoding="utf-8") as f:
        f.write(xml_declaration)
        tree.write(f, encoding="unicode")
    
    print(f"Successfully converted {excel_path} to {output_iprj_path}")

def dxf_to_excel(dxf_path, output_excel_path, bg_width=1600):
    """
    Convert DXF file to Excel format suitable for IPRJ conversion
    
    Args:
        dxf_path: Path to the DXF file
        output_excel_path: Path to save the output Excel file
    """
    print(f"Reading DXF file: {dxf_path}")
    
    # Load the DXF file
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    
    # Find background rectangle and determine origin and scaling
    origin_x, origin_y, meter_per_pixel, scale_factor = find_background_rectangle(msp, bg_width)
    print(f"Using origin at: ({origin_x}, {origin_y}) survey feet")
    print(f"Meter per pixel: {meter_per_pixel}, Scale factor: {scale_factor}")
    
    # Process background image based on background rectangle
    background_data = process_background_image(msp, dxf_path, origin_x, origin_y, meter_per_pixel)
    
    # Process sensors and their event zones with the new scaling
    sensor_data = process_sensors(msp, origin_x, origin_y, scale_factor)
    
    # Generate Excel workbook
    create_excel_file(sensor_data, background_data, output_excel_path)
    
    print(f"Successfully converted {dxf_path} to {output_excel_path}")
    
def find_background_rectangle(msp, bg_width=1600):
    """Find the background rectangle to determine origin and scaling"""
    # Find entities in the "Background" layer
    background_entities = [entity for entity in msp if entity.dxf.layer == 'Background']
    
    print(f"Found {len(background_entities)} entities in Background layer")
    
    # Look for various types of entities that could be rectangles
    rectangles = []
    
    for entity in background_entities:
        entity_type = entity.dxftype()

        if entity_type == '3DFACE':
            # LWPOLYLINEs can be rectangles
            points = list(entity.wcs_vertices())
            if len(points) == 4 or (len(points) == 5 and np.allclose(points[0], points[4])):
                rectangles.append(('3DFACE', entity, points))
                print(f"  Found 3DFACE with {len(points)} points")        
 
        if entity_type == 'LWPOLYLINE':
            # LWPOLYLINEs can be rectangles
            points = list(entity.get_points())
            if len(points) == 4 or (len(points) == 5 and np.allclose(points[0], points[4])):
                rectangles.append(('LWPOLYLINE', entity, points))
                print(f"  Found LWPOLYLINE with {len(points)} points")
        
        elif entity_type == 'POLYLINE':
            # Regular POLYLINEs can be rectangles too
            vertices = list(entity.vertices)
            points = [v.dxf.location for v in vertices]
            if len(points) == 4 or (len(points) == 5 and np.allclose(points[0], points[4])):
                rectangles.append(('POLYLINE', entity, points))
                print(f"  Found POLYLINE with {len(points)} points")
    
    if not rectangles:
        print("Warning: No background rectangle found. Using default values.")
        # Check all entities for a potential bounding box fallback
        all_points = []
        for entity in msp:
            if hasattr(entity, 'get_points'):
                try:
                    all_points.extend(list(entity.get_points()))
                except:
                    pass
            elif entity.dxftype() == 'CIRCLE':
                cx, cy = entity.dxf.center.x, entity.dxf.center.y
                r = entity.dxf.radius
                all_points.extend([(cx-r, cy-r), (cx+r, cy+r)])
        
        if all_points:
            # Use the bounding box of all entities as fallback
            x_coords = [p[0] for p in all_points]
            y_coords = [p[1] for p in all_points]
            origin_x, origin_y = min(x_coords), min(y_coords)
            width = max(x_coords) - min(x_coords)
            print(f"Using fallback bounding box with width {width} feet")
            scale_factor = bg_width / width
            meter_per_pixel = width * 0.3048 / bg_width
            return origin_x, origin_y, meter_per_pixel, scale_factor
        
        return 0, 0, 0.3048, 1.0  # Default values if all else fails
    
    # Use the first rectangle found
    rect_type, rect_entity, points = rectangles[0]
    print(f"Using {rect_type} as background rectangle")
    
    # Calculate rectangle properties
    x_coords = [p[0] for p in points]
    y_coords = [p[1] for p in points]
    
    # Origin is the bottom-left corner
    origin_x = min(x_coords)
    origin_y = max(y_coords)
    
    # Calculate width in feet
    width_feet = max(x_coords) - min(x_coords)
    
    # Calculate meter per pixel and scale factor
    # Rectangle width will be mapped to background width in pixels
    scale_factor = bg_width / width_feet  # pixels per foot
    meter_per_pixel = width_feet * 0.3048 / bg_width  # meters per pixel
    
    return origin_x, origin_y, meter_per_pixel, scale_factor

def process_background_image(msp, dxf_path, origin_x, origin_y, meter_per_pixel):
    """Process background image references based on background rectangle"""
    # Create image path by changing extension from dxf to jpg
    image_path = os.path.splitext(dxf_path)[0] + ".jpg"
    
    # Create background data dictionary
    background_data = {
        "image_path": image_path,
        "Background_PosX": "0",  # Origin in pixels (top-left corner of rectangle)
        "Background_PosY": "0",
        "BackgroundImageRotation": "0",  # No rotation
        "MeterPerPixel": str(meter_per_pixel),  # New field for meter per pixel ratio
    }
    
    return background_data

def process_sensors(msp, origin_x, origin_y, scale_factor):
    """Process sensors and their event zones with new scaling and data structure"""
    sensor_data = {}

    # Look for sensor layers (Sensor1, Sensor2, etc.)
    for sensor_idx in range(1, 5):  # Support up to 4 sensors
        layer_name = f"Sensor{sensor_idx}"

        # Check if this sensor layer exists
        sensor_entities = [entity for entity in msp if entity.dxf.layer == layer_name]

        if not sensor_entities:
            continue  # Skip if no entities in this sensor layer

        print(f"Processing {len(sensor_entities)} entities in {layer_name} layer")

        # Initialize data structure for this sensor
        sensor_data[sensor_idx - 1] = {
            "positional": {},
            "eventzones": [],
            "ignorezones": [],
            "conditions": []
        }

        # Find sensor position (circle)
        sensor_circles = [entity for entity in sensor_entities if entity.dxftype() == 'CIRCLE']

        if sensor_circles:
            circle = sensor_circles[0]  # Use the first circle as the sensor
            print(f"Found sensor circle at ({circle.dxf.center.x}, {circle.dxf.center.y})")
            
            # Convert coordinates to pixels relative to new origin
            # This translates the coordinate so that origin_x, origin_y becomes pixel 0,0
            sensor_x_pixels = (circle.dxf.center.x - origin_x) * scale_factor
            sensor_y_pixels = (origin_y - circle.dxf.center.y) * scale_factor

            print(f"Converted to pixel coordinates ({sensor_x_pixels}, {sensor_y_pixels})")

            # Store sensor positional data
            sensor_data[sensor_idx - 1]["positional"]["Position_X"] = sensor_x_pixels
            sensor_data[sensor_idx - 1]["positional"]["Position_Y"] = sensor_y_pixels
            sensor_data[sensor_idx - 1]["positional"]["AzimuthAngle"] = 0  # Default values
            sensor_data[sensor_idx - 1]["positional"]["ElevationAngle"] = 0
            sensor_data[sensor_idx - 1]["positional"]["RoadGradientAngle"] = 0
            sensor_data[sensor_idx - 1]["positional"]["InstallationHeight"] = 6  # Default height in meters

        # Find event zones (polygons)
        polygon_entity_types = ('LWPOLYLINE', 'POLYLINE', 'HATCH', '3DFACE')
        polygons = [entity for entity in sensor_entities if entity.dxftype() in polygon_entity_types]

        print(f"Found {len(polygons)} potential event zone polygons")

        # Sort polygons by transparency in reverse order
        polygons.sort(key=lambda p: getattr(p.dxf, 'transparency', 0) if hasattr(p.dxf, 'transparency') else 0, reverse=True)

        # Process each polygon as an event zone
        for zone_idx, polygon in enumerate(polygons):
            entity_type = polygon.dxftype()
            print(f"Processing event zone {zone_idx+1} of type {entity_type}")
            
            # Get polygon points based on entity type
            points = []
            if entity_type == 'LWPOLYLINE':
                points = list(polygon.get_points())
            elif entity_type == 'POLYLINE':
                points = [vertex.dxf.location for vertex in polygon.vertices]
            elif entity_type == 'HATCH':
                try:
                    # Use the first boundary path
                    path = polygon.paths[0]
                    points = path.vertices
                except (AttributeError, IndexError):
                    continue  # Skip if we can't get points
            elif entity_type == '3DFACE':
                points = list(polygon.wcs_vertices())

            if not points:
                print(f"Skipping event zone {zone_idx+1} - no points found")
                continue
                
            # Number of zone points
            nr_of_zone_points = len(points)
            print(f"Event zone has {nr_of_zone_points} points")
            
            # Get color and map to PhaseNumber
            color = getattr(polygon.dxf, 'color', 0) if hasattr(polygon.dxf, 'color') else 0
            phase_number = color if 1 <= color <= 16 else 0
            
            # Basic zone data with updated field names
            lineweight_mapping = {0: 0, 13: 1, 30: 2}
            zone_type = lineweight_mapping.get(
                getattr(polygon.dxf, 'lineweight', 0), 0
            ) if hasattr(polygon.dxf, 'lineweight') else 0

            zone_data = {
                "Index": zone_idx+1,
                "Enable": 1,
                "ZoneName": f"Ph {phase_number}",
                "NrOfZonePoints": nr_of_zone_points,
                "ZoneType": zone_type,
                "PhaseNumber": phase_number,
                "EventMessageDelay": 0,
                "EventMessageExtend": 0,
                "OutputNumber": 0,
                "EtaEnable": 0,
                "EtaPoint_X": 0,
                "EtaPoint_Y": 0
            }

            # Convert coordinates relative to new origin and scale to pixels
            for i, point in enumerate(points):
                # Extract x, y based on point type
                if isinstance(point, tuple):
                    # LWPOLYLINE points are tuples
                    x_dxf, y_dxf = point[0], point[1]
                elif hasattr(point, 'x') and hasattr(point, 'y'):
                    # ezdxf.math.Vec2 or similar
                    x_dxf, y_dxf = point.x, point.y
                else:
                    # Assume it's some kind of sequence
                    try:
                        x_dxf, y_dxf = point[0], point[1]
                    except (IndexError, TypeError):
                        print(f"Couldn't extract coordinates from point: {point}")
                        continue
                
                # Convert to pixels relative to the background rectangle origin
                x_pixels = (x_dxf - origin_x) * scale_factor
                y_pixels = (origin_y - y_dxf) * scale_factor

                zone_data[f"ZonePoint_{i}_X"] = x_pixels
                zone_data[f"ZonePoint_{i}_Y"] = y_pixels

            # Add zone to the sensor's event zones
            sensor_data[sensor_idx - 1]["eventzones"].append(zone_data)

    return sensor_data

def create_excel_file(sensor_data, background_data, output_excel_path):
    """Create Excel file with the processed data"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Function to add a section title
    def add_section_title(worksheet, row, title):
        cell = worksheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True)
        return row + 1
    
    # Process each sensor
    for sensor_idx, sensor in sensor_data.items():
        ws = wb.create_sheet(title=f"Sensor {sensor_idx + 1}")
        current_row = 1
        
        # Write positional data vertically
        pos_items = list(sensor["positional"].items())
        for i, (key, val) in enumerate(pos_items):
            ws[f"A{current_row + i}"] = key
            ws[f"B{current_row + i}"] = val
        
        current_row += len(pos_items) + 2  # Add empty row
        
        # Write Event Zones section
        current_row = add_section_title(ws, current_row, "Event Zones")
        
        # Find the maximum number of points across all zones
        max_points = 0
        for zone in sensor["eventzones"]:
            points_count = zone.get("NrOfZonePoints", 0)
            max_points = max(max_points, points_count)
        
        # Create event zones dataframe
        df_zones = pd.DataFrame(sensor["eventzones"])
    
        if not df_zones.empty:
            # Rename "Index" column to "EventZone"
            df_zones.rename(columns={"Index": "EventZone"}, inplace=True)

            # Ensure all possible ZonePoint columns exist
            for i in range(max_points):
                if f"ZonePoint_{i}_X" not in df_zones.columns:
                    df_zones[f"ZonePoint_{i}_X"] = np.nan
                    df_zones[f"ZonePoint_{i}_Y"] = np.nan

            # Sort columns to match required order
            base_columns = ["EventZone", "Enable", "ZoneName", "NrOfZonePoints", "ZoneType", "PhaseNumber", 
                            "EventMessageDelay", "EventMessageExtend", "OutputNumber", "EtaEnable", 
                            "EtaPoint_X", "EtaPoint_Y"]

            point_columns = []
            for i in range(max_points):
                point_columns.extend([f"ZonePoint_{i}_X", f"ZonePoint_{i}_Y"])

            # Combine and reorder columns
            all_columns = base_columns + point_columns
            all_columns = [col for col in all_columns if col in df_zones.columns]
            df_zones = df_zones[all_columns]

            # Write dataframe to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df_zones, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)

            current_row += len(df_zones) + 2  # Include header row and add empty row
        else:
            # Add empty header row with the new column names
            headers = ["EventZone", "Enable", "ZoneName", "NrOfZonePoints", "ZoneType", "PhaseNumber", 
                    "EventMessageDelay", "EventMessageExtend", "OutputNumber", "EtaEnable", 
                    "EtaPoint_X", "EtaPoint_Y", "ZonePoint_0_X", "ZonePoint_0_Y", 
                    "ZonePoint_1_X", "ZonePoint_1_Y", "ZonePoint_2_X", "ZonePoint_2_Y",
                    "ZonePoint_3_X", "ZonePoint_3_Y"]

            for c_idx, header in enumerate(headers):
                ws.cell(row=current_row, column=c_idx + 1, value=header)
            current_row += 2  # Header row + empty row
        
        # Write Conditions section (empty for now)
        current_row = add_section_title(ws, current_row, "Conditions")
        headers = ["EventZone", "Enable", "OutputNumber", "ConditionClass",
                  "Direction", "VelocityMin", "VelocityMax"]
        for c_idx, header in enumerate(headers):
            ws.cell(row=current_row, column=c_idx + 1, value=header)
        current_row += 2  # Header row + empty row
        
        # Write Ignore Zones section (empty for now)
        current_row = add_section_title(ws, current_row, "Ignore Zones")
        headers = ["Index", "Enable", "ZonePoint_0_X", "ZonePoint_0_Y", 
                  "ZonePoint_1_X", "ZonePoint_1_Y", "ZonePoint_2_X", "ZonePoint_2_Y",
                  "ZonePoint_3_X", "ZonePoint_3_Y"]
        for c_idx, header in enumerate(headers):
            ws.cell(row=current_row, column=c_idx + 1, value=header)
    
    # Create Annotations sheet (empty for now)
    ws_annotations = wb.create_sheet("Annotations")
    current_row = 1
    
    # Add Lineals section
    current_row = add_section_title(ws_annotations, current_row, "Lineals")
    headers = ["Index", "Type", "Enable", "Point_0_X", "Point_0_Y", "Point_1_X", "Point_1_Y"]
    for c_idx, header in enumerate(headers):
        ws_annotations.cell(row=current_row, column=c_idx + 1, value=header)
    current_row += 2  # Header row + empty row
    
    # Add Textlabels section
    current_row = add_section_title(ws_annotations, current_row, "Textlabels")
    headers = ["Index", "Type", "Enable", "Text", "Position_X", "Position_Y", "FontSize"]
    for c_idx, header in enumerate(headers):
        ws_annotations.cell(row=current_row, column=c_idx + 1, value=header)
    
    # Create Background sheet if we have data
    if background_data:
        ws_bg = wb.create_sheet("BackgroundImageInfo")
        
        # Write background data vertically
        for i, (key, val) in enumerate(background_data.items()):
            ws_bg[f"A{i+1}"] = key
            ws_bg[f"B{i+1}"] = val
    
    wb.save(output_excel_path)